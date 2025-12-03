import unicodedata
import re
from decimal import Decimal, InvalidOperation
from datetime import datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Q

import openpyxl

# Asegúrate de importar tus modelos
from nomina.models import (
    Cargo, Condicion, EscalaSalarial, Empleado, DatosBancarios, Quincena,
    AsignacionesMensuales, AsignacionesQuincenales, Deducciones,
    AsignacionAdicionalMensual, AsignacionAdicionalQuincenal, DeduccionAdicional
)


# --------------------------
# Utilidades de normalización
# --------------------------
def normalize_text(s):
    """Normaliza un texto: None->'', strip, lower, quitar acentos, compactar espacios."""
    if s is None:
        return ''
    if not isinstance(s, str):
        s = str(s)
    s = s.strip()
    # reemplazar varios espacios por uno solo
    s = re.sub(r'\s+', ' ', s)
    # bajar a minusculas
    s = s.lower()
    # quitar acentos
    s = unicodedata.normalize('NFKD', s)
    s = ''.join([c for c in s if not unicodedata.combining(c)])
    s = s.strip()
    return s


def col_letter(col_idx):
    """Devuelve letra de columna (1->A, 2->B, ...)"""
    # openpyxl tiene utility pero evitamos importarlo adicionalmente
    result = ''
    while col_idx:
        col_idx, rem = divmod(col_idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def parse_decimal(value):
    """Intenta parsear a Decimal. Si vacio -> Decimal('0')."""
    if value is None:
        return Decimal('0')
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return Decimal('0')
    s = str(value).strip()
    if s == '':
        return Decimal('0')
    # quitar símbolos no numéricos comunes
    s = s.replace(',', '')  # si usan 1,234.56 o 1234, necesitarías adaptar
    s = re.sub(r'[^\d\.-]', '', s)
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal('0')


def parse_int(value):
    if value is None:
        return None
    if isinstance(value, int):
        return value
    try:
        s = str(value).strip()
        if s == '':
            return None
        return int(float(s))
    except Exception:
        return None


def parse_date(value):
    """Intenta convertir a date. Puede venir como datetime.date, datetime, o string dd/mm/yyyy."""
    if value is None or value == '':
        return None
    if hasattr(value, 'date') and hasattr(value, 'year'):
        # datetime.date or datetime.datetime
        try:
            return value.date() if hasattr(value, 'date') else value
        except Exception:
            pass
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    # fallback: intenta parse int timestamp
    try:
        ts = float(s)
        return datetime.fromtimestamp(ts).date()
    except Exception:
        return None


# --------------------------
# Búsquedas flexibles de encabezados
# --------------------------

# Encabezados oficiales de cada rango
COLUMNAS_BASE = {
    "mensual": [
        "ESCALA SALARIAL",
        "SUELDO BASE MENSUAL",
        "PRIMA DE ANTIGÜEDAD",
        "PRIMA DE PROFESIONALIZACION",
        "PRIMA POR HIJOS",
        "CONTRIBUCION PARA TRABAJADORAS Y TRABAJADORES CON DISCAPACIDAD",
        "HORAS EXTRAS",
        "COMPLEMENTO DEL SALARIO",
        "BECAS PARA HIJOS",
        "PRIMA ASISTENCIAL Y DEL HOGAR",
        "PRIMA TRABAJADORES ADM Y OBR",
        "ENCARGADURIA",
        "TOTAL MENSUAL",
        "TOTAL ASIGNACION MENSUAL (TODOS LOS CONCEPTOS)",
    ],
    "quincenal": [
        "SUELDO BASE QUINCENAL",
        "PRIMA DE ANTIGÜEDAD",
        "PRIMA DE PROFESIONALIZACION",
        "PRIMA POR HIJOS",
        "CONTRIBUCION PARA TRABAJADORAS Y TRABAJADORES CON DISCAPACIDAD",
        "COMPLEMENTO DEL SALARIO",
        "BECAS PARA HIJOS",
        "PRIMA ASISTENCIAL Y DEL HOGAR",
        "PRIMA TRABAJADORES ADM Y OBR",
        "ENCARGADURIA",
        "DIFERENCIA POR COMISIONES DE SERVICIOS",
        "RETROACTIVO SUELDO BASE",
        "RETROACTIVO PRIMA DE ANTIGÜEDAD",
        "RETROACTIVO PRIMA DE PROFESIONALIZACION",
        "RETROACTIVO PRIMA POR HIJOS",
        "RETROACTIVO CONTRIBUCION PARA TRABAJADORAS Y TRABAJADORES CON DISCAPACIDAD",
        "RETROACTIVO HORAS EXTRAS",
        "RETROACTIVO BECAS PARA HIJOS",
        "RETROACTIVO PRIMA ASISTENCIAL Y DEL HOGAR",
        "RETROACTIVO PRIMA TRABAJADORES ADM Y OBR",
        "RETROACTIVO ENCARGADURIA",
        "TOTAL DE ASIGNACION QUINCENAL (TODOS LOS CONCEPTOS)",
    ],
    "deduccion": [
        "RETENCION POR C.AHORRO 15%",
        "RETENCION POR S.S.O",
        "RETENCION R.P.E",
        "RETENCION POR F.A.O.V",
        "RETENCION POR F.E.J.P",
        "SINAEP (SINDICATO NACIONAL DE EMPLEADOS PUBLICOS DEL MPPE)",
        "IPASME (INSTITUTO DE PREVENSION Y ASISTENCIA SOCIAL)",
        "DESCUENTO POR PAGO INDEBIDO",
        "TOTAL DEDUCCIONES",
    ],
}

# Diccionarios de nombres esperados (normalizados) -> Lista de variantes
HEADER_VARIANTS = {
    # claves para tablas y campos
    'cedula': ['cedula', 'cédula', 'ci'],
    'cargo': ['cargo'],
    'condicion': ['condicion', 'condición'],
    'escala_salarial': ['escala salarial', 'escala', 'nivel'],
    'nombres_y_apellidos': ['nombres y apellidos', 'nombre', 'nombre completo', 'nombres'],
    'unidad_de_adscripcion_y_o_direccion': ['unidad de adscripcion y o direccion', 'unidad de adscripción', 'unidad', 'dirección'],
    'estado_o_ubicacion': ['estado o ubicacion', 'estado', 'ubicacion'],
    'codigo_del_cargo': ['codigo del cargo', 'código del cargo', 'codigo cargo'],
    'fecha_de_ingreso': ['fecha de ingreso', 'fecha ingreso'],
    'grado_de_instruccion': ['grado de instruccion', 'grado instruccion', 'grado de instrucción'],
    'anos_previos': ['años previos', 'anos previos', 'años anteriores'],
    'anos_en_fundabit': ['años en fundabit', 'anos en fundabit'],
    'total_anos_de_servicios': ['total años de servicios', 'total anos de servicios'],
    'numero_de_hijos': ['nº de hijos', 'numero de hijos', 'número de hijos'],
    'discapacitado': ['contribucion para trabajadoras y trabajadores con discapacidad', 'discapacitado', 'contribucion discapacidad'],
    'especialidad': ['especialidad'],
    'sexo': ['sexo', 'genero'],
    'fecha_de_nacimiento': ['fecha de nacimiento', 'nacimiento'],

    # Datos bancarios
    'banco': ['banco'],
    'numero_de_cuenta': ['numero de cuenta', 'número de cuenta', 'cuenta'],

    # Quincena (varios encabezados)
    'dias_laborados': ['dias laborados', 'días laborados', 'dias trabajados'],
    'horas_extras_quincena': ['horas extras', 'horas extras quincena'],
    'total_retroactivo': ['total retroactivo'],
    'total_asignacion_primera_quincena': ['total asignación 1era quincena', 'total asignacion 1era quincena', 'total asignacion primera quincena', 'total asignacion 1era'],
    'total_asignacion_segunda_quincena': ['total asignacion 2da quincena', 'total asignacion segunda quincena', 'total asignacion 2°', 'total asignacion 2° quincena', 'total asignacion 2da'],
    'total_de_asignacion_quincenal_todos_los_conceptos': ['total de asignacion quincenal (todos los conceptos)', 'total de asignacion quincenal', 'total de asignacion quincenal todos los conceptos'],
    'total_mensual': ['total mensual', 'total mensual (todos los conceptos)'],
    'total_asignacion_mensual_todos_los_conceptos': ['total asignacion mensual (todos los conceptos)', 'total asignacion mensual'],
    'total_deducciones': ['total deducciones', 'total de deducciones'],
    'total_a_cancelar': ['total a cancelar', 'total a cobrar', 'a cancelar'],
    'nota': ['nota', 'observacion', 'observación'],

    # Asignaciones Mensuales (ejemplos)
    'sueldo_base_mensual': ['sueldo base mensual', 'sueldo mensual'],
    'prima_de_antiguedad': ['prima de antiguedad', 'prima antiguedad'],
    'prima_de_profesionalizacion': ['prima de profesionalizacion', 'prima profesionalizacion'],
    'prima_por_hijos': ['prima por hijos', 'prima hijos'],
    'contribucion_para_trabajadoras_y_trabajadores_con_discapacidad_mensual': ['contribucion para trabajadoras y trabajadores con discapacidad'],
    'horas_extras_mensual': ['horas extras'],
    'complemento_del_salario_mensual': ['complemento del salario'],
    'becas_para_hijos_mensual': ['becas para hijos', 'becas hijos'],
    'prima_asistencial_y_del_hogar_mensual': ['prima asistencial y del hogar'],
    'prima_trabajadores_adm_y_obr_mensual': ['prima trabajadores adm y obr'],
    'encargaduria_mensual': ['encargaduria'],

    # Asignaciones Quincenales (ejemplos)
    'sueldo_base_quincenal': ['sueldo base quincenal'],
    'prima_de_antiguedad_quincenal': ['prima de antiguedad'],
    'prima_de_profesionalizacion_quincenal': ['prima de profesionalizacion'],
    'prima_por_hijos_quincenal': ['prima por hijos'],
    'contribucion_para_trabajadoras_y_trabajadores_con_discapacidad_quincenal': ['contribucion para trabajadoras y trabajadores con discapacidad'],
    'complemento_del_salario_quincenal': ['complemento del salario'],
    'becas_para_hijos_quincenal': ['becas para hijos'],
    'prima_asistencial_y_del_hogar_quincenal': ['prima asistencial y del hogar'],
    'prima_trabajadores_adm_y_obr_quincenal': ['prima trabajadores adm y obr'],
    'encargaduria_quincenal': ['encargaduria'],
    'diferencia_por_comisiones_de_servicios': ['diferencia por comisiones de servicios'],
    'retroactivo_sueldo_base': ['retroactivo sueldo base'],
    'retroactivo_prima_de_antiguedad': ['retroactivo prima de antiguedad'],
    'retroactivo_prima_de_profesionalizacion': ['retroactivo prima de profesionalizacion'],
    'retroactivo_prima_por_hijos': ['retroactivo prima por hijos'],
    'retroactivo_contribucion_para_trabajadoras_y_trabajadores_con_discapacidad': ['retroactivo contribucion discapacidad'],
    'retroactivo_horas_extras': ['retroactivo horas extras'],
    'retroactivo_becas_para_hijos': ['retroactivo becas para hijos'],
    'retroactivo_prima_asistencial_y_del_hogar': ['retroactivo prima asistencial y del hogar'],
    'retroactivo_prima_trabajadores_adm_y_obr': ['retroactivo prima trabajadores adm y obr'],
    'retroactivo_encargaduria': ['retroactivo encargaduria'],

    # Deducciones
    'retencion_por_caja_de_ahorro': ['retencion por c.ahorro 15%', 'retencion por caja de ahorro', 'retencion por c.ahorro'],
    'retencion_por_sso': ['retencion por s.s.o', 'retencion por sso'],
    'retencion_rpe': ['retencion r.p.e', 'retencion rpe'],
    'retencion_por_faov': ['retencion por f.a.o.v', 'retencion por faov'],
    'retencion_por_fejp': ['retencion por f.e.j.p', 'retencion por fejp'],
    'sinaep': ['sinaep'],
    'ipasme': ['ipasme'],
    'descuento_por_pago_indebido': ['descuento por pago indebido'],
}

# invertimos HEADER_VARIANTS para hacer búsqueda: variante_normalizada -> key
VARIANT_MAP = {}
for key, variants in HEADER_VARIANTS.items():
    for v in variants:
        VARIANT_MAP[normalize_text(v)] = key


def header_to_key(header_norm):
    """Mapea un encabezado normalizado a la 'key' de HEADER_VARIANTS si es posible."""
    if not header_norm:
        return None
    if header_norm in VARIANT_MAP:
        return VARIANT_MAP[header_norm]
    # búsqueda por tokens: si alguno de los tokens coincide fuertemente
    tokens = header_norm.split()
    for t in tokens:
        if t in VARIANT_MAP:
            return VARIANT_MAP[t]
    # heurística: contiene ciertas palabras clave
    heuristics = {
        'cedula': ['cedula', 'cédula', 'ci'],
        'cargo': ['cargo'],
        'condicion': ['condicion'],
        'escala_salarial': ['escala', 'nivel'],
        'nombres_y_apellidos': ['nombre', 'apellido'],
        'banco': ['banco'],
        'numero_de_cuenta': ['cuenta'],
        'dias_laborados': ['dias', 'labor'],
        'total_deducciones': ['deducc'],
        'total_a_cancelar': ['cancelar', 'a cancelar'],
        'total_mensual': ['mes', 'mensual'],
        # etc. Esto ayuda si la cabecera no está exactamente.
    }
    for k, keys in heuristics.items():
        for token in keys:
            if token in header_norm:
                return k
    return None


# --------------------------
# Función principal del comando
# --------------------------
class Command(BaseCommand):
    help = "Importa quincenas desde un archivo .xlsx siguiendo la lógica especificada."

    def add_arguments(self, parser):
        # No argumentos CLI: interactivo
        pass

    def handle(self, *args, **options):
        self.stdout.write(self.style.NOTICE("IMPORTAR QUINCENA - comando interactivo"))
        periodo = None
        while periodo not in (1, 2):
            try:
                periodo = int(input("Ingrese periodo (1=Primera quincena, 2=Segunda quincena): ").strip())
            except Exception:
                periodo = None

        mes = None
        while mes is None or not (1 <= mes <= 12):
            try:
                mes = int(input("Ingrese mes (1-12): ").strip())
            except Exception:
                mes = None

        ano = None
        while ano is None:
            try:
                ano = int(input("Ingrese año (ej. 2025): ").strip())
            except Exception:
                ano = None

        # pedir ruta del archivo
        archivo_path = None
        while not archivo_path:
            ruta = input("Ingrese la ruta completa al archivo .xlsx: ").strip()
            if ruta:
                p = Path(ruta)
                if p.exists() and p.suffix.lower() in ('.xlsx',):
                    archivo_path = str(p)
                else:
                    self.stdout.write(self.style.ERROR("Ruta inválida o no es .xlsx, inténtalo de nuevo."))
        self.stdout.write(self.style.NOTICE(f"Periodo: {periodo} - Mes: {mes} - Año: {ano}"))
        self.stdout.write(self.style.NOTICE(f"Archivo: {archivo_path}"))

        # Validar contra últimos periodos en Quincena
        # Orden por (ano, mes, periodo) asc -> último es mayor
        # Obtén el último registro (si existe)
        ultimas_quincenas = list(Quincena.objects.order_by('ano', 'mes', 'periodo').values('ano', 'mes', 'periodo').distinct())
        ultimo = None
        if ultimas_quincenas:
            ultimo = ultimas_quincenas[-1]
            ultimo_key = (int(ultimo['ano']), int(ultimo['mes']), int(ultimo['periodo']))
            input_key = (ano, mes, periodo)
            if input_key < ultimo_key:
                raise CommandError(f"El periodo a cargar ({periodo}/{mes}/{ano}) es anterior al último cargado ({ultimo['periodo']}/{ultimo['mes']}/{ultimo['ano']}). Abortando.")
            else:
                if input_key == ultimo_key:
                    self.stdout.write(self.style.WARNING("El periodo ingresado es IGUAL al último registrado -> se permitirá sobreescritura (corrección)."))
                    sobreescribir_igual = True
                else:
                    sobreescribir_igual = False
        else:
            self.stdout.write(self.style.NOTICE("No se encontraron registros previos de quincenas. Se continuará normalmente."))
            sobreescribir_igual = False

        # Cargar workbook
        wb = openpyxl.load_workbook(archivo_path, data_only=True)
        # Asumimos la primera hoja
        ws = wb.worksheets[0]

        # Analizar fila 1 para obtener encabezados y sus columnas
        header_row_idx = 1
        max_col = ws.max_column
        headers = {}  # col_idx -> original_text
        headers_norm = {}  # col_idx -> normalized
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            val = cell.value
            if val is not None and str(val).strip() != '':
                headers[col_idx] = str(val).strip()
                headers_norm[col_idx] = normalize_text(str(val))

        # Mapeo encabezado normalizado -> col index
        mapped_headers = {}  # key (HEADER_VARIANTS key) -> col_idx
        # También guardar reverse map: col_idx -> mapped_key (si se pudo)
        col_to_key = {}

        for col_idx, norm in headers_norm.items():
            key = header_to_key(norm)
            if key:
                # solo si no asignado ya, preferimos primera aparición
                if key not in mapped_headers:
                    mapped_headers[key] = col_idx
                    col_to_key[col_idx] = key

        # Determinar los anclajes para rangos:
        # Buscamos los encabezados exactos de anclaje (normalizados)
        anchor_names = [
            'escala salarial',
            'total asignacion mensual (todos los conceptos)',
            'total de asignacion quincenal (todos los conceptos)',
            'total deducciones',
        ]
        # normalizamos anchors
        anchors_norm = [normalize_text(x) for x in anchor_names]
        anchor_positions = {}
        for col_idx, norm in headers_norm.items():
            if norm in anchors_norm:
                anchor_positions[norm] = col_idx

        # Si no encontramos usando texto exacto, intentamos heurística:
        # buscar el header que contenga palabras clave
        for key_anchor in anchors_norm:
            if key_anchor not in anchor_positions:
                for col_idx, norm in headers_norm.items():
                    if key_anchor.split()[0] in norm:
                        anchor_positions.setdefault(key_anchor, col_idx)

        # Asignar rangos por índice (inclusive inicio, inclusive fin)
        # Rangos solicitados por el usuario:
        # 1: ESCALA SALARIAL -> TOTAL ASIGNACION MENSUAL (TODOS LOS CONCEPTOS)
        # 2: TOTAL ASIGNACION MENSUAL (TODOS LOS CONCEPTOS) -> TOTAL DE ASIGNACION QUINCENAL (TODOS LOS CONCEPTOS)
        # 3: TOTAL DE ASIGNACION QUINCENAL (TODOS LOS CONCEPTOS) -> TOTAL DEDUCCIONES
        # Nota: si algún ancla no se encuentra, asumimos rangos extremos.
        def find_anchor_col_by_text(possible_texts):
            for t in possible_texts:
                t_norm = normalize_text(t)
                if t_norm in anchor_positions:
                    return anchor_positions[t_norm]
            return None

        col_escala = find_anchor_col_by_text(['escala salarial'])
        col_total_asig_mensual_anchor = find_anchor_col_by_text(['total asignacion mensual (todos los conceptos)', 'total mensual'])
        col_total_asig_quincenal_anchor = find_anchor_col_by_text(['total de asignacion quincenal (todos los conceptos)', 'total de asignacion quincenal'])
        col_total_deducciones_anchor = find_anchor_col_by_text(['total deducciones', 'total de deducciones'])

        # Si no logra ubicar alguno de los anclajes, intenta ubicar por heurística por contenido
        if not col_escala:
            for col_idx, norm in headers_norm.items():
                if 'escala' in norm or 'nivel' in norm:
                    col_escala = col_idx
                    break
        if not col_total_asig_mensual_anchor:
            for col_idx, norm in headers_norm.items():
                if 'total mensual' in norm or 'total asignacion mensual' in norm:
                    col_total_asig_mensual_anchor = col_idx
                    break
        if not col_total_asig_quincenal_anchor:
            for col_idx, norm in headers_norm.items():
                if 'total de asignacion quincenal' in norm or 'total asignacion quincenal' in norm:
                    col_total_asig_quincenal_anchor = col_idx
                    break
        if not col_total_deducciones_anchor:
            for col_idx, norm in headers_norm.items():
                if 'deduccion' in norm or 'deducciones' in norm:
                    col_total_deducciones_anchor = col_idx
                    break

        # Si aun no encuentra, setea por limites:
        if not col_escala:
            col_escala = 1
        if not col_total_asig_mensual_anchor:
            col_total_asig_mensual_anchor = col_escala + 5 if (col_escala + 5) <= max_col else col_escala
        if not col_total_asig_quincenal_anchor:
            col_total_asig_quincenal_anchor = col_total_asig_mensual_anchor + 5 if (col_total_asig_mensual_anchor + 5) <= max_col else col_total_asig_mensual_anchor
        if not col_total_deducciones_anchor:
            col_total_deducciones_anchor = col_total_asig_quincenal_anchor + 5 if (col_total_asig_quincenal_anchor + 5) <= max_col else col_total_asig_quincenal_anchor

        # Rango 1: col_escala .. col_total_asig_mensual_anchor (inclusive)
        rango1_start = col_escala
        rango1_end = col_total_asig_mensual_anchor

        # Rango 2: col_total_asig_mensual_anchor .. col_total_asig_quincenal_anchor
        rango2_start = col_total_asig_mensual_anchor
        rango2_end = col_total_asig_quincenal_anchor

        # Rango 3: col_total_asig_quincenal_anchor .. col_total_deducciones_anchor
        rango3_start = col_total_asig_quincenal_anchor
        rango3_end = col_total_deducciones_anchor

        self.stdout.write(self.style.NOTICE(f"Rango Asignaciones Mensuales: columnas {rango1_start} ({col_letter(rango1_start)}) a {rango1_end} ({col_letter(rango1_end)})"))
        self.stdout.write(self.style.NOTICE(f"Rango Asignaciones Quincenales: columnas {rango2_start} ({col_letter(rango2_start)}) a {rango2_end} ({col_letter(rango2_end)})"))
        self.stdout.write(self.style.NOTICE(f"Rango Deducciones: columnas {rango3_start} ({col_letter(rango3_start)}) a {rango3_end} ({col_letter(rango3_end)})"))

        # -------------------------
        # Registrar Cargos, Condicion, EscalaSalarial (valores únicos)
        # -------------------------
        # Buscar columna "CARGO"
        cargo_col = mapped_headers.get('cargo') or None
        if cargo_col:
            valores = set()
            for row in range(2, ws.max_row + 1):
                v = ws.cell(row=row, column=cargo_col).value
                if v is None:
                    continue
                vn = str(v).strip()
                if vn == '':
                    continue
                valores.add(vn)
            # comparar con DB
            existentes = set(Cargo.objects.values_list('cargo', flat=True))
            nuevos = [v for v in valores if v not in existentes]
            for nv in nuevos:
                Cargo.objects.create(cargo=nv)
            self.stdout.write(self.style.SUCCESS(f"Tabla Cargos: {len(nuevos)} nuevos cargados."))

        # Condicion
        condicion_col = mapped_headers.get('condicion') or None
        if condicion_col:
            valores = set()
            for row in range(2, ws.max_row + 1):
                v = ws.cell(row=row, column=condicion_col).value
                if v is None:
                    continue
                vn = str(v).strip()
                if vn == '':
                    continue
                valores.add(vn)
            existentes = set(Condicion.objects.values_list('condicion', flat=True))
            nuevos = [v for v in valores if v not in existentes]
            for nv in nuevos:
                Condicion.objects.create(condicion=nv)
            self.stdout.write(self.style.SUCCESS(f"Tabla Condicion: {len(nuevos)} nuevos cargados."))

        # EscalaSalarial (col "ESCALA SALARIAL")
        escala_col = None
        # intenta encontrar por mapped_headers
        for k in ('escala_salarial',):
            if k in mapped_headers:
                escala_col = mapped_headers[k]
        # si no, intenta heurística
        if not escala_col:
            for col_idx, norm in headers_norm.items():
                if 'escala' in norm or 'nivel' in norm:
                    escala_col = col_idx
                    break
        if escala_col:
            valores = set()
            for row in range(2, ws.max_row + 1):
                v = ws.cell(row=row, column=escala_col).value
                if v is None:
                    continue
                vn = str(v).strip()
                if vn == '':
                    continue
                valores.add(vn)
            existentes = set(EscalaSalarial.objects.values_list('nivel', flat=True))
            nuevos = [v for v in valores if v not in existentes]
            for nv in nuevos:
                EscalaSalarial.objects.create(nivel=nv)
            self.stdout.write(self.style.SUCCESS(f"Tabla EscalaSalarial: {len(nuevos)} nuevos cargados."))

        # -------------------------
        # Preparar mapeo de columnas por 'key' usando encabezados detectados
        # -------------------------
        # Mapeo exacto: key -> col_idx (ya parcialmente hecho en mapped_headers)
        # Para campos con múltiples variantes que pueden compartir columnas, usamos header_to_key heurístico.
        # Ya tenemos 'mapped_headers' con las coincidencias "fuertes".

        # -------------------------
        # Verificar si existen registros para el mismo periodo: en tal caso borrar relacionados
        # -------------------------
        periodo_qs = Quincena.objects.filter(periodo=periodo, mes=mes, ano=ano)
        if periodo_qs.exists():
            # sacar los ids de esas quincenas
            ids = list(periodo_qs.values_list('id', flat=True))
            # borrar AsignacionesMensuales, AsignacionesQuincenales, Deducciones que referencien esos ids
            AsignacionesMensuales.objects.filter(quincena_id__in=ids).delete()
            AsignacionesQuincenales.objects.filter(quincena_id__in=ids).delete()
            Deducciones.objects.filter(quincena_id__in=ids).delete()

            # AsignacionAdicionalQuincenal, AsignacionAdicionalMensual, DeduccionAdicional -> quitar quincenas y eliminar si ya no tienen quincenas

            AsignacionAdicionalMensual.objects.filter(quincena_id__in=ids).delete()
            AsignacionAdicionalQuincenal.objects.filter(quincena_id__in=ids).delete()
            DeduccionAdicional.objects.filter(quincena_id__in=ids).delete()

            # finalmente borrar las quincenas
            periodo_qs.delete()
            self.stdout.write(self.style.WARNING(f"Se eliminaron registros anteriores del periodo {periodo}/{mes}/{ano} y sus relaciones."))
        else:
            self.stdout.write(self.style.NOTICE("No existían registros previos para el periodo indicado."))

        # -------------------------
        # Recorrer filas y procesar cada empleado (cada CEDULA)
        # -------------------------
        # localizar columna CEDULA
        cedula_col = None
        for col_idx, norm in headers_norm.items():
            if normalize_text(norm) in ('cedula', 'cédula', 'ci') or 'cedula' in norm:
                cedula_col = col_idx
                break
        if not cedula_col:
            # intentar mapped_headers 'cedula'
            cedula_col = mapped_headers.get('cedula')
        if not cedula_col:
            raise CommandError("No se encontró la columna 'CEDULA' en el encabezado. Abortando.")

        filas_procesadas = 0
        with transaction.atomic():
            for row in range(2, ws.max_row + 1):
                cedula_val = ws.cell(row=row, column=cedula_col).value
                if cedula_val is None:
                    continue
                try:
                    cedula_int = int(str(cedula_val).strip())
                except Exception:
                    # si la cédula es no numérica, saltar
                    continue
                # Comenzamos a construir/actualizar empleado
                # Primero, obtener valores de columnas referenciadas
                # Cargo
                cargo_val = None
                if cargo_col:
                    cargo_val = ws.cell(row=row, column=cargo_col).value
                    cargo_val = str(cargo_val).strip() if cargo_val is not None else None
                # Condicion
                condicion_val = None
                if condicion_col:
                    condicion_val = ws.cell(row=row, column=condicion_col).value
                    condicion_val = str(condicion_val).strip() if condicion_val is not None else None
                # Escala
                escala_val = None
                if escala_col:
                    escala_val = ws.cell(row=row, column=escala_col).value
                    escala_val = str(escala_val).strip() if escala_val is not None else None

                # Otros campos: vamos a buscar usando header->col map (mapped_headers)
                def get_cell_by_key(key):
                    col = mapped_headers.get(key)
                    if not col:
                        # intentar buscar en headers_norm por heurística
                        for cidx, hn in headers_norm.items():
                            if key.replace('_', ' ') in hn:
                                return ws.cell(row=row, column=cidx).value
                        return None
                    return ws.cell(row=row, column=col).value

                nombres_val = get_cell_by_key('nombres_y_apellidos') or ''
                unidad_val = get_cell_by_key('unidad_de_adscripcion_y_o_direccion') or ''
                estado_val = get_cell_by_key('estado_o_ubicacion') or ''
                codigo_cargo_val = get_cell_by_key('codigo_del_cargo') or ''
                fecha_ingreso_val = parse_date(get_cell_by_key('fecha_de_ingreso'))
                grado_instruccion_val = get_cell_by_key('grado_de_instruccion') or ''
                anos_previos_val = parse_int(get_cell_by_key('anos_previos')) or 0
                anos_fundabit_val = parse_int(get_cell_by_key('anos_en_fundabit')) or 0
                total_anos_servicios_val = parse_int(get_cell_by_key('total_anos_de_servicios')) or 0
                num_hijos_val = parse_int(get_cell_by_key('numero_de_hijos')) or 0
                discap_val_raw = get_cell_by_key('discapacitado') or get_cell_by_key('contribucion_para_trabajadoras_y_trabajadores_con_discapacidad')
                discap_val = False
                if discap_val_raw is not None:
                    try:
                        if str(discap_val_raw).strip() in ('1', '1.0', 'true', 't', 'yes', 'si', 'sí'):
                            discap_val = True
                    except Exception:
                        discap_val = False
                especialidad_val = get_cell_by_key('especialidad') or ''
                sexo_val = get_cell_by_key('sexo') or ''
                fecha_nac_val = parse_date(get_cell_by_key('fecha_de_nacimiento'))

                # Datos bancarios
                banco_val = get_cell_by_key('banco') or ''
                num_cuenta_val = get_cell_by_key('numero_de_cuenta') or get_cell_by_key('cuenta') or ''

                # Buscar empleado por cedula
                empleado_qs = Empleado.objects.filter(cedula=cedula_int)
                es_creacion = not empleado_qs.exists()

                # Resolver FKs: Cargo, Condicion, EscalaSalarial
                cargo_obj = None
                if cargo_val:
                    cargo_obj, created = Cargo.objects.get_or_create(cargo=cargo_val)

                condicion_obj = None
                if condicion_val:
                    condicion_obj, created = Condicion.objects.get_or_create(condicion=condicion_val)

                escala_obj = None
                if escala_val:
                    escala_obj, created = EscalaSalarial.objects.get_or_create(nivel=escala_val)

                if es_creacion:
                    # Crear nuevo empleado
                    empleado = Empleado.objects.create(
                        cargo=cargo_obj,
                        condicion=condicion_obj,
                        escala_salarial=escala_obj,
                        cedula=cedula_int,
                        nombres_y_apellidos=str(nombres_val)[:255] if nombres_val else '',
                        unidad_de_adscripcion_y_o_direccion=str(unidad_val)[:200] if unidad_val else '',
                        estado_o_ubicacion=str(estado_val)[:100] if estado_val else '',
                        codigo_del_cargo=str(codigo_cargo_val)[:50] if codigo_cargo_val else None,
                        fecha_de_ingreso=fecha_ingreso_val or datetime.today().date(),
                        grado_de_instruccion=str(grado_instruccion_val)[:100] if grado_instruccion_val else '',
                        anos_previos=anos_previos_val,
                        anos_en_fundabit=anos_fundabit_val,
                        total_anos_de_servicios=total_anos_servicios_val,
                        numero_de_hijos=num_hijos_val,
                        discapacitado=discap_val,
                        especialidad=str(especialidad_val)[:100] if especialidad_val else None,
                        sexo=str(sexo_val)[:20] if sexo_val else '',
                        fecha_de_nacimiento=fecha_nac_val or datetime.today().date(),
                    )
                    # Datos bancarios
                    if banco_val or num_cuenta_val:
                        try:
                            numero_cuenta_int = int(str(num_cuenta_val).strip()) if str(num_cuenta_val).strip() != '' else None
                        except Exception:
                            numero_cuenta_int = None
                        if numero_cuenta_int is not None:
                            DatosBancarios.objects.create(empleado=empleado, banco=str(banco_val)[:100], numero_de_cuenta=numero_cuenta_int)
                else:
                    # Existe -> actualizar si necesario
                    empleado = empleado_qs.first()
                    # Si estamos sobreescribiendo (periodo igual al ultimo), hacemos update() directo para que
                    # django-simple-history NO tome esto como nuevo historial: se actualiza en DB sin crear historial.
                    if sobreescribir_igual:
                        update_fields = {}
                        if cargo_obj and (empleado.cargo_id != cargo_obj.id):
                            update_fields['cargo_id'] = cargo_obj.id
                        if condicion_obj and (empleado.condicion_id != condicion_obj.id):
                            update_fields['condicion_id'] = condicion_obj.id
                        if escala_obj and (empleado.escala_salarial_id != escala_obj.id):
                            update_fields['escala_salarial_id'] = escala_obj.id

                        if nombres_val and empleado.nombres_y_apellidos != nombres_val:
                            update_fields['nombres_y_apellidos'] = str(nombres_val)[:255]
                        if unidad_val and empleado.unidad_de_adscripcion_y_o_direccion != unidad_val:
                            update_fields['unidad_de_adscripcion_y_o_direccion'] = str(unidad_val)[:200]
                        if estado_val and empleado.estado_o_ubicacion != estado_val:
                            update_fields['estado_o_ubicacion'] = str(estado_val)[:100]
                        if codigo_cargo_val and empleado.codigo_del_cargo != codigo_cargo_val:
                            update_fields['codigo_del_cargo'] = str(codigo_cargo_val)[:50]
                        if fecha_ingreso_val and empleado.fecha_de_ingreso != fecha_ingreso_val:
                            update_fields['fecha_de_ingreso'] = fecha_ingreso_val
                        if grado_instruccion_val and empleado.grado_de_instruccion != grado_instruccion_val:
                            update_fields['grado_de_instruccion'] = str(grado_instruccion_val)[:100]
                        if anos_previos_val is not None and empleado.anos_previos != anos_previos_val:
                            update_fields['anos_previos'] = anos_previos_val
                        if anos_fundabit_val is not None and empleado.anos_en_fundabit != anos_fundabit_val:
                            update_fields['anos_en_fundabit'] = anos_fundabit_val
                        if total_anos_servicios_val is not None and empleado.total_anos_de_servicios != total_anos_servicios_val:
                            update_fields['total_anos_de_servicios'] = total_anos_servicios_val
                        if num_hijos_val is not None and empleado.numero_de_hijos != num_hijos_val:
                            update_fields['numero_de_hijos'] = num_hijos_val
                        if empleado.discapacitado != discap_val:
                            update_fields['discapacitado'] = discap_val
                        if especialidad_val and ((empleado.especialidad or '') != especialidad_val):
                            update_fields['especialidad'] = str(especialidad_val)[:100]
                        if sexo_val and empleado.sexo != sexo_val:
                            update_fields['sexo'] = str(sexo_val)[:20]
                        if fecha_nac_val and empleado.fecha_de_nacimiento != fecha_nac_val:
                            update_fields['fecha_de_nacimiento'] = fecha_nac_val

                        if update_fields:
                            Empleado.objects.filter(pk=empleado.pk).update(**update_fields)
                            # refresh empleado
                            empleado = Empleado.objects.get(pk=empleado.pk)
                    else:
                        # Periodo nuevo: permitimos save normal (storico)
                        changed = False
                        if cargo_obj and (empleado.cargo_id != cargo_obj.id):
                            empleado.cargo = cargo_obj
                            changed = True
                        if condicion_obj and (empleado.condicion_id != condicion_obj.id):
                            empleado.condicion = condicion_obj
                            changed = True
                        if escala_obj and (empleado.escala_salarial_id != escala_obj.id):
                            empleado.escala_salarial = escala_obj
                            changed = True
                        if nombres_val and empleado.nombres_y_apellidos != nombres_val:
                            empleado.nombres_y_apellidos = str(nombres_val)[:255]; changed = True
                        if unidad_val and empleado.unidad_de_adscripcion_y_o_direccion != unidad_val:
                            empleado.unidad_de_adscripcion_y_o_direccion = str(unidad_val)[:200]; changed = True
                        if estado_val and empleado.estado_o_ubicacion != estado_val:
                            empleado.estado_o_ubicacion = str(estado_val)[:100]; changed = True
                        if codigo_cargo_val and empleado.codigo_del_cargo != codigo_cargo_val:
                            empleado.codigo_del_cargo = str(codigo_cargo_val)[:50]; changed = True
                        if fecha_ingreso_val and empleado.fecha_de_ingreso != fecha_ingreso_val:
                            empleado.fecha_de_ingreso = fecha_ingreso_val; changed = True
                        if grado_instruccion_val and empleado.grado_de_instruccion != grado_instruccion_val:
                            empleado.grado_de_instruccion = str(grado_instruccion_val)[:100]; changed = True
                        if anos_previos_val is not None and empleado.anos_previos != anos_previos_val:
                            empleado.anos_previos = anos_previos_val; changed = True
                        if anos_fundabit_val is not None and empleado.anos_en_fundabit != anos_fundabit_val:
                            empleado.anos_en_fundabit = anos_fundabit_val; changed = True
                        if total_anos_servicios_val is not None and empleado.total_anos_de_servicios != total_anos_servicios_val:
                            empleado.total_anos_de_servicios = total_anos_servicios_val; changed = True
                        if num_hijos_val is not None and empleado.numero_de_hijos != num_hijos_val:
                            empleado.numero_de_hijos = num_hijos_val; changed = True
                        if empleado.discapacitado != discap_val:
                            empleado.discapacitado = discap_val; changed = True
                        if especialidad_val and ((empleado.especialidad or '') != especialidad_val):
                            empleado.especialidad = str(especialidad_val)[:100]; changed = True
                        if sexo_val and empleado.sexo != sexo_val:
                            empleado.sexo = str(sexo_val)[:20]; changed = True
                        if fecha_nac_val and empleado.fecha_de_nacimiento != fecha_nac_val:
                            empleado.fecha_de_nacimiento = fecha_nac_val; changed = True
                        if changed:
                            empleado.save()  # esto creará histórico si simple_history está activo

                # Datos Bancarios: OneToOne con empleado
                # si sobreescribir_igual: usar update; si no, usar save/create
                if banco_val or num_cuenta_val:
                    try:
                        numero_cuenta_int = int(str(num_cuenta_val).strip()) if str(num_cuenta_val).strip() != '' else None
                    except Exception:
                        numero_cuenta_int = None
                    if hasattr(empleado, 'datos_bancarios') and empleado.datos_bancarios is not None:
                        # existe
                        if sobreescribir_igual:
                            update_fields = {}
                            if banco_val and (empleado.datos_bancarios.banco != banco_val):
                                update_fields['banco'] = str(banco_val)[:100]
                            if numero_cuenta_int is not None and (empleado.datos_bancarios.numero_de_cuenta != numero_cuenta_int):
                                update_fields['numero_de_cuenta'] = numero_cuenta_int
                            if update_fields:
                                DatosBancarios.objects.filter(pk=empleado.datos_bancarios.pk).update(**update_fields)
                        else:
                            changed = False
                            db = empleado.datos_bancarios
                            if banco_val and db.banco != banco_val:
                                db.banco = str(banco_val)[:100]; changed = True
                            if numero_cuenta_int is not None and db.numero_de_cuenta != numero_cuenta_int:
                                db.numero_de_cuenta = numero_cuenta_int; changed = True
                            if changed:
                                db.save()
                    else:
                        # no existe -> crear si hay datos
                        if numero_cuenta_int is not None:
                            DatosBancarios.objects.create(
                                empleado=empleado,
                                banco=str(banco_val)[:100] if banco_val else '',
                                numero_de_cuenta=numero_cuenta_int
                            )

                # -------------------------
                # Crear Quincena y tablas relacionadas
                # -------------------------
                # quincena (campo) no está usado para validar; usamos periodo/mes/ano
                quincena_obj = Quincena.objects.create(
                    empleado=empleado,
                    periodo=periodo,
                    mes=mes,
                    ano=ano,
                    fecha=datetime.today().date(),
                    dias_laborados=parse_int(get_cell_by_key('dias_laborados')) or 0,
                    horas_extras_quincena=parse_decimal(ws.cell(row=row, column=mapped_headers.get('horas_extras_quincena', mapped_headers.get('horas_extras', None))).value) if mapped_headers.get('horas_extras_quincena') or mapped_headers.get('horas_extras') else Decimal('0'),
                    total_retroactivo=parse_decimal(get_cell_by_key('total_retroactivo')),
                    total_asignacion_primera_quincena=parse_decimal(get_cell_by_key('total_asignacion_primera_quincena')),
                    total_asignacion_segunda_quincena=parse_decimal(get_cell_by_key('total_asignacion_segunda_quincena')),
                    total_de_asignacion_quincenal_todos_los_conceptos=parse_decimal(get_cell_by_key('total_de_asignacion_quincenal_todos_los_conceptos')),
                    total_mensual=parse_decimal(get_cell_by_key('total_mensual')),
                    total_asignacion_mensual_todos_los_conceptos=parse_decimal(get_cell_by_key('total_asignacion_mensual_todos_los_conceptos')),
                    total_deducciones=parse_decimal(get_cell_by_key('total_deducciones')),
                    total_a_cancelar=parse_decimal(get_cell_by_key('total_a_cancelar')),
                    nota=str(get_cell_by_key('nota') or '')[:2000]
                )

                # AsignacionesMensuales (OneToOne)
                am = AsignacionesMensuales.objects.create(
                    quincena=quincena_obj,
                    sueldo_base_mensual=parse_decimal(get_cell_by_key('sueldo_base_mensual')),
                    prima_de_antiguedad=parse_decimal(get_cell_by_key('prima_de_antiguedad')),
                    prima_de_profesionalizacion=parse_decimal(get_cell_by_key('prima_de_profesionalizacion')),
                    prima_por_hijos=parse_decimal(get_cell_by_key('prima_por_hijos')),
                    contribucion_para_trabajadoras_y_trabajadores_con_discapacidad=parse_decimal(get_cell_by_key('contribucion_para_trabajadoras_y_trabajadores_con_discapacidad')),
                    horas_extras=parse_decimal(get_cell_by_key('horas_extras')),
                    complemento_del_salario=parse_decimal(get_cell_by_key('complemento_del_salario')),
                    becas_para_hijos=parse_decimal(get_cell_by_key('becas_para_hijos')),
                    prima_asistencial_y_del_hogar=parse_decimal(get_cell_by_key('prima_asistencial_y_del_hogar')),
                    prima_trabajadores_adm_y_obr=parse_decimal(get_cell_by_key('prima_trabajadores_adm_y_obr')),
                    encargaduria=parse_decimal(get_cell_by_key('encargaduria')),
                )

                # AsignacionesQuincenales
                aq = AsignacionesQuincenales.objects.create(
                    quincena=quincena_obj,
                    sueldo_base_quincenal=parse_decimal(get_cell_by_key('sueldo_base_quincenal')),
                    prima_de_antiguedad_quincenal=parse_decimal(get_cell_by_key('prima_de_antiguedad_quincenal')),
                    prima_de_profesionalizacion_quincenal=parse_decimal(get_cell_by_key('prima_de_profesionalizacion_quincenal')),
                    prima_por_hijos_quincenal=parse_decimal(get_cell_by_key('prima_por_hijos_quincenal')),
                    contribucion_para_trabajadoras_y_trabajadores_con_discapacidad_quincenal=parse_decimal(get_cell_by_key('contribucion_para_trabajadoras_y_trabajadores_con_discapacidad_quincenal')),
                    complemento_del_salario_quincenal=parse_decimal(get_cell_by_key('complemento_del_salario_quincenal')),
                    becas_para_hijos_quincenal=parse_decimal(get_cell_by_key('becas_para_hijos_quincenal')),
                    prima_asistencial_y_del_hogar_quincenal=parse_decimal(get_cell_by_key('prima_asistencial_y_del_hogar_quincenal')),
                    prima_trabajadores_adm_y_obr_quincenal=parse_decimal(get_cell_by_key('prima_trabajadores_adm_y_obr_quincenal')),
                    encargaduria_quincenal=parse_decimal(get_cell_by_key('encargaduria_quincenal')),
                    diferencia_por_comisiones_de_servicios=parse_decimal(get_cell_by_key('diferencia_por_comisiones_de_servicios')),
                    retroactivo_sueldo_base=parse_decimal(get_cell_by_key('retroactivo_sueldo_base')),
                    retroactivo_prima_de_antiguedad=parse_decimal(get_cell_by_key('retroactivo_prima_de_antiguedad')),
                    retroactivo_prima_de_profesionalizacion=parse_decimal(get_cell_by_key('retroactivo_prima_de_profesionalizacion')),
                    retroactivo_prima_por_hijos=parse_decimal(get_cell_by_key('retroactivo_prima_por_hijos')),
                    retroactivo_contribucion_para_trabajadoras_y_trabajadores_con_discapacidad=parse_decimal(get_cell_by_key('retroactivo_contribucion_para_trabajadoras_y_trabajadores_con_discapacidad')),
                    retroactivo_horas_extras=parse_decimal(get_cell_by_key('retroactivo_horas_extras')),
                    retroactivo_becas_para_hijos=parse_decimal(get_cell_by_key('retroactivo_becas_para_hijos')),
                    retroactivo_prima_asistencial_y_del_hogar=parse_decimal(get_cell_by_key('retroactivo_prima_asistencial_y_del_hogar')),
                    retroactivo_prima_trabajadores_adm_y_obr=parse_decimal(get_cell_by_key('retroactivo_prima_trabajadores_adm_y_obr')),
                    retroactivo_encargaduria=parse_decimal(get_cell_by_key('retroactivo_encargaduria')),
                )

                # Deducciones
                ded = Deducciones.objects.create(
                    quincena=quincena_obj,
                    retencion_por_caja_de_ahorro=parse_decimal(get_cell_by_key('retencion_por_caja_de_ahorro')),
                    retencion_por_sso=parse_decimal(get_cell_by_key('retencion_por_sso')),
                    retencion_rpe=parse_decimal(get_cell_by_key('retencion_rpe')),
                    retencion_por_faov=parse_decimal(get_cell_by_key('retencion_por_faov')),
                    retencion_por_fejp=parse_decimal(get_cell_by_key('retencion_por_fejp')),
                    sinaep=parse_decimal(get_cell_by_key('sinaep')),
                    ipasme=parse_decimal(get_cell_by_key('ipasme')),
                    descuento_por_pago_indebido=parse_decimal(get_cell_by_key('descuento_por_pago_indebido')),
                )

                # -------------------------
                # Asignaciones/Deducciones adicionales (rango dinámico)
                # -------------------------
                # Rango 1 (Mensuales): recorrer encabezados en rango1_start..rango1_end
                for c in range(rango1_start + 1, rango1_end - 1):
                    if c in headers:
                            header_text = headers[c]
                            if normalize_text(header_text) in [normalize_text(h) for h in COLUMNAS_BASE["mensual"]]:
                                continue  # columna estándar, saltar

                            valor_raw = ws.cell(row=row, column=c).value
                            if valor_raw is None:
                                continue
                            valor_dec = parse_decimal(valor_raw)
                            if valor_dec != Decimal('0'):
                                a = AsignacionAdicionalMensual.objects.create(
                                    quincena=quincena_obj,
                                    nombre=str(header_text)[:100],
                                    valor=valor_dec
                                )

                # Rango 2 (Quincenales)
                for c in range(rango2_start + 2, rango2_end - 1):
                    if c in headers:
                        header_text = headers[c]
                        if normalize_text(header_text) in [normalize_text(h) for h in COLUMNAS_BASE["quincenal"]]:
                            continue  # columna estándar, saltar

                        valor_raw = ws.cell(row=row, column=c).value
                        if valor_raw is None:
                            continue
                        valor_dec = parse_decimal(valor_raw)
                        if valor_dec != Decimal('0'):
                            a = AsignacionAdicionalQuincenal.objects.create(
                                quincena=quincena_obj,
                                nombre=str(header_text)[:100],
                                valor=valor_dec
                            )

                # Rango 3 (Deducciones)
                for c in range(rango3_start + 1, rango3_end - 1):
                    if c in headers:
                        header_text = headers[c]
                        if normalize_text(header_text) in [normalize_text(h) for h in COLUMNAS_BASE["deduccion"]]:
                            continue  # columna estándar, saltar

                        valor_raw = ws.cell(row=row, column=c).value
                        if valor_raw is None:
                            continue
                        valor_dec = parse_decimal(valor_raw)
                        if valor_dec != Decimal('0'):
                            d = DeduccionAdicional.objects.create(
                                quincena=quincena_obj,
                                nombre=str(header_text)[:100],
                                valor=valor_dec
                            )

                filas_procesadas += 1

        self.stdout.write(self.style.SUCCESS(f"Importación finalizada. Filas procesadas: {filas_procesadas}"))
        self.stdout.write(self.style.SUCCESS("Listo."))
